import os
import json
import pickle
import csv
import io
import traceback
from datetime import datetime, date, timedelta
from flask import Flask, request, redirect, url_for, render_template, jsonify

import openpyxl

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "rg-backorder-2026")

DATA_DIR = os.environ.get("DATA_DIR", "/tmp/backorder_data")
os.makedirs(DATA_DIR, exist_ok=True)

PARTS_FILE   = os.path.join(DATA_DIR, "parts.pkl")
NOTES_FILE   = os.path.join(DATA_DIR, "notes.json")
FEEDER_FILE  = os.path.join(DATA_DIR, "feeder.json")
PLANNER_FILE = os.path.join(DATA_DIR, "planner.json")
META_FILE    = os.path.join(DATA_DIR, "meta.json")
ERROR_FILE   = os.path.join(DATA_DIR, "last_error.txt")

PURCHASED_PLANNERS = {"boes","craig","devowe","glynn","salcedo","slifer","zhang","tracy"}


# ── Storage ───────────────────────────────────────────────────

def load_notes():
    try:
        with open(NOTES_FILE) as f: return json.load(f)
    except: return {}

def save_notes(notes):
    with open(NOTES_FILE,"w") as f: json.dump(notes, f)

def load_feeder():
    try:
        with open(FEEDER_FILE) as f: return set(json.load(f))
    except: return None

def save_feeder(s):
    with open(FEEDER_FILE,"w") as f: json.dump(list(s), f)

def load_parts():
    try:
        with open(PARTS_FILE,"rb") as f: return pickle.load(f)
    except: return None

def save_parts(data):
    with open(PARTS_FILE,"wb") as f: pickle.dump(data, f)

def load_meta():
    try:
        with open(META_FILE) as f: return json.load(f)
    except: return {}

def save_meta(m):
    with open(META_FILE,"w") as f: json.dump(m, f)

def load_planner():
    try:
        with open(PLANNER_FILE) as f: return json.load(f)
    except: return {}

def save_planner(d):
    with open(PLANNER_FILE,"w") as f: json.dump(d, f)

def save_error(msg):
    try:
        with open(ERROR_FILE,"w") as f: f.write(msg)
    except: pass


# ── Parsing ───────────────────────────────────────────────────

def fmt_date(val):
    if not val or str(val).strip() in ("", "None", "nan"): return ""
    s = str(val).strip()
    # Strip time component: "5/29/2026 12:00:00 AM" -> "5/29/2026"
    s = s.split(' ')[0].strip()
    # Try parsing the full date string (handles single-digit months/days)
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%#m/%#d/%Y", "%-m/%-d/%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except: pass
    # Fallback: try first 10 chars for ISO-style dates
    for fmt in ("%Y-%m-%d",):
        try: return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except: pass
    # Excel serial number
    try:
        serial = int(float(s))
        if 40000 < serial < 60000:
            d = datetime(1899, 12, 30) + timedelta(days=serial)
            return d.strftime("%Y-%m-%d")
    except: pass
    return ""

def col_find(headers, *names):
    lu = {str(h).lower().replace(" ","").replace("_","").replace("#",""): i
          for i, h in enumerate(headers)}
    for n in names:
        k = n.lower().replace(" ","").replace("_","").replace("#","")
        if k in lu: return lu[k]
    return None

def safe_float(val):
    try: return float(str(val).replace(",","").strip())
    except: return 0.0

def read_xlsx_rows(file_bytes):
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), read_only=True, data_only=True, keep_links=False
    )
    best_ws, best_cols = wb.worksheets[0], 0
    for ws in wb.worksheets:
        first = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))
        ncols = sum(1 for c in first if c is not None)
        if ncols > best_cols:
            best_cols = ncols
            best_ws = ws
    all_rows = list(best_ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows: return [], []
    headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    return headers, all_rows[1:]

def read_csv_rows(file_bytes):
    for enc in ("utf-8-sig", "latin-1", "utf-8"):
        try:
            text = file_bytes.decode(enc)
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if rows: return rows[0], rows[1:]
        except: pass
    return [], []

def process_rows(headers, rows, feeder_set):
    C = {
        "coord":   col_find(headers,"project coordinator","coordinator"),
        "proj":    col_find(headers,"customer address","customeraddress","project"),
        "order":   col_find(headers,"order no","orderno","order"),
        "po_no":   col_find(headers,"po no","pono","po#","ponumber"),
        "po_st":   col_find(headers,"po status","postatus"),
        "part":    col_find(headers,"part no","partno","part","partnumber"),
        "line_st": col_find(headers,"line status","linestatus"),
        "oqty":    col_find(headers,"order qty","orderqty"),
        "sqty":    col_find(headers,"shipped qty","shippedqty"),
        "opqty":   col_find(headers,"open qty","openqty"),
        "inv":     col_find(headers,"inventory"),
        "val":     col_find(headers,"unshipped value","unshippedvalue"),
        "ship":    col_find(headers,"ship date","shipdate"),
        "due":     col_find(headers,"due date","duedate"),
        "bldg":    col_find(headers,"building"),
        "planner": col_find(headers,"planner","buyer"),
        "purch":   col_find(headers,"purchased"),
    }

    def g(row, key):
        idx = C.get(key)
        if idx is None or idx >= len(row): return ""
        v = row[idx]
        return "" if v is None else str(v).strip()

    today = date.today().strftime("%Y-%m-%d")
    parts, summary, max_date, row_count = {}, {}, "", 0
    _planner_cache = load_planner()  # cache once before loop

    for row in rows:
        if not any(c for c in row if c is not None and str(c).strip()):
            continue
        row_count += 1
        coord   = g(row,"coord") or "Unassigned"
        proj    = g(row,"proj")  or "Unknown"
        key     = f"{coord}|||{proj}"
        raw_part = g(row,"part")
        # Strip trailing ' - SUFFIX' (e.g. 'H6BFG1NGT - ' or 'A18MIE0003633 - G')
        part_no = raw_part.split(' - ')[0].strip().rstrip('- ').strip()
        oqty    = safe_float(g(row,"oqty"))
        sqty    = safe_float(g(row,"sqty"))
        opqty   = safe_float(g(row,"opqty"))
        inv     = safe_float(g(row,"inv"))
        val     = safe_float(g(row,"val"))
        ship    = fmt_date(g(row,"ship"))
        due     = fmt_date(g(row,"due"))
        overdue = 1 if ship and "2020" < ship < "2100" and ship < today else 0
        if ship and "2020" < ship < "2100" and ship > max_date:
            max_date = ship
        purchased = 0
        purch_val = g(row,"purch").strip()
        if purch_val:
            # Report has a Purchased column — use it as primary source (case-insensitive)
            purchased = 1 if purch_val.lower() == "yes" else 0
        elif feeder_set is not None:
            # No Purchased column — fall back to feeder set matching
            purchased = 1 if part_no in feeder_set else 0
        # Get planner from row if column exists, otherwise look up from feeder
        row_planner = g(row,"planner")
        if not row_planner:
            row_planner = _planner_cache.get(part_no, '')
        p = [g(row,"order"), g(row,"po_st"), part_no, g(row,"line_st"),
             oqty, sqty, opqty, inv, val, ship, due, overdue,
             g(row,"bldg"), purchased, row_planner, g(row,"po_no")]
        parts.setdefault(key, []).append(p)
        summary.setdefault(coord, {})
        summary[coord].setdefault(proj, {"open_qty":0,"inventory":0,"value":0,"count":0,"overdue":0})
        m = summary[coord][proj]
        m["open_qty"] += opqty; m["inventory"] += inv; m["value"] += val; m["count"] += 1
        if overdue: m["overdue"] += 1

    return parts, summary, today, row_count  # use upload date, not max ship date

def parse_one_feeder_csv(file_bytes):
    """Parse a single feeder CSV and return (set of purchased part numbers, part->planner dict)."""
    purchased = set()
    planner_map = {}
    for enc in ("utf-8-sig","latin-1","utf-8"):
        try:
            text = file_bytes.decode(enc)
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                src        = row.get("Part Source","").lower()
                planner    = row.get("Planner","").strip()
                part_no    = row.get("Part No","").strip()
                if not part_no: continue
                # Store planner mapping
                if planner and part_no not in planner_map:
                    planner_map[part_no] = planner
                    base = part_no.split(':')[0].strip()
                    if base != part_no and base not in planner_map:
                        planner_map[base] = planner
                # Purchased logic
                is_purch = ("purchased" in src) or ("manufactured" in src and planner.lower() in PURCHASED_PLANNERS)
                if is_purch:
                    purchased.add(part_no)
                    base = part_no.split(':')[0].strip()
                    if base != part_no:
                        purchased.add(base)
            return purchased, planner_map
        except: continue
    return purchased, {}


# ── Routes ────────────────────────────────────────────────────

@app.route("/", methods=["GET"])
def index():
    feeder = load_feeder()
    return render_template("index.html",
        meta=load_meta(),
        has_data=load_parts() is not None,
        has_feeder=feeder is not None,
        feeder_size=len(feeder) if feeder else 0,
        note_count=len(load_notes()),
        error=None,
    )

@app.route("/upload", methods=["POST"])
def upload():
    report_file = request.files.get("report")
    if not report_file or not report_file.filename:
        return redirect(url_for("index"))
    try:
        raw   = report_file.read()
        fname = report_file.filename.lower()
        if fname.endswith(".csv"):
            headers, rows = read_csv_rows(raw)
        else:
            headers, rows = read_xlsx_rows(raw)
        if not headers:
            raise ValueError("No headers found — check the file format.")
        feeder_set = load_feeder()
        parts, summary, report_date, row_count = process_rows(headers, rows, feeder_set)
        if not parts:
            raise ValueError(f"No data rows parsed. Headers: {headers[:6]}")
        save_parts({"parts": parts, "summary": summary})
        save_meta({
            "report_date":  report_date,
            "report_name":  report_file.filename,
            "row_count":    row_count,
            "uploaded_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
        })
        return redirect(url_for("dashboard"))
    except Exception as e:
        save_error(traceback.format_exc())
        return render_template("index.html",
            meta=load_meta(), has_data=load_parts() is not None,
            has_feeder=load_feeder() is not None,
            feeder_size=len(load_feeder()) if load_feeder() else 0,
            note_count=len(load_notes()), error=str(e))


@app.route("/upload-feeder", methods=["POST"])
def upload_feeder():
    """Upload ONE feeder CSV at a time — merges into saved feeder set."""
    f = request.files.get("feeder_csv")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "No file received"})
    try:
        raw = f.read()
        new_parts, new_planner = parse_one_feeder_csv(raw)
        existing = load_feeder() or set()
        merged = existing | new_parts
        save_feeder(merged)
        # Merge planner map
        existing_planner = load_planner()
        existing_planner.update(new_planner)
        save_planner(existing_planner)
        # Re-tag existing backorder data and update planners
        data = load_parts()
        if data:
            for key, part_list in data["parts"].items():
                for p in part_list:
                    pv = p[2]
                    p[13] = 1 if pv in merged else 0
                    if not p[14]:  # only fill planner if missing
                        p[14] = existing_planner.get(pv, '')
            save_parts(data)
        return jsonify({
            "ok": True,
            "filename": f.filename,
            "this_file_parts": len(new_parts),
            "cumulative_total": len(merged),
            "previously_had": len(existing),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()})

@app.route("/feeder-status")
def feeder_status():
    """Check what feeder data is saved on the server."""
    feeder = load_feeder()
    data = load_parts()
    purchased = 0
    total = 0
    if data:
        for pl in data["parts"].values():
            for p in pl:
                total += 1
                if p[13]: purchased += 1
    return jsonify({
        "feeder_loaded": feeder is not None,
        "feeder_size": len(feeder) if feeder else 0,
        "expected_size": "~337000",
        "feeder_ok": len(feeder) > 100000 if feeder else False,
        "backorder_total": total,
        "backorder_purchased": purchased,
        "meta": load_meta(),
    })

@app.route("/clear-feeder", methods=["POST"])
def clear_feeder_route():
    try: os.remove(FEEDER_FILE)
    except: pass
    return jsonify({"ok": True})





@app.route("/dashboard")
def dashboard():
    data = load_parts()
    if not data: return redirect(url_for("index"))
    feeder = load_feeder()
    return render_template("dashboard.html",
        meta=load_meta(),
        parts_json=json.dumps(data["parts"],   separators=(",",":")),
        summary_json=json.dumps(data["summary"], separators=(",",":")),
        notes_json=json.dumps(load_notes(),    separators=(",",":")),
        has_feeder=feeder is not None,
        feeder_size=len(feeder) if feeder else 0,
    )

@app.route("/api/notes", methods=["GET"])
def get_notes():
    return jsonify(load_notes())

@app.route("/api/notes", methods=["POST"])
def set_notes():
    save_notes(request.get_json(force=True) or {})
    return jsonify({"ok": True})

@app.route("/test-purchased")
def test_purchased():
    """Shows exactly what purchased logic the server is running."""
    data = load_parts()
    if not data:
        return jsonify({"error": "No data loaded yet"})
    
    total = sum(len(v) for v in data["parts"].values())
    purchased = sum(p[13] for pl in data["parts"].values() for p in pl)
    
    # Sample purchased parts
    purch_samples = []
    not_purch_samples = []
    for key, pl in data["parts"].items():
        for p in pl:
            if p[13] and len(purch_samples) < 5:
                purch_samples.append({"part": p[2], "proj": key.split("|||")[1]})
            if not p[13] and len(not_purch_samples) < 5:
                not_purch_samples.append({"part": p[2], "proj": key.split("|||")[1]})
    
    feeder = load_feeder()
    return jsonify({
        "total_parts": total,
        "purchased_count": purchased,
        "purchased_pct": round(purchased/total*100, 1) if total else 0,
        "feeder_loaded": feeder is not None,
        "feeder_size": len(feeder) if feeder else 0,
        "purchased_samples": purch_samples,
        "not_purchased_samples": not_purch_samples,
        "meta": load_meta(),
    })

@app.route("/debug")
def debug():
    data = load_parts(); feeder = load_feeder()
    try: err = open(ERROR_FILE).read()
    except: err = ""
    return jsonify({
        "meta": load_meta(),
        "has_parts": data is not None,
        "parts_keys": list(data["parts"].keys())[:5] if data else [],
        "has_feeder": feeder is not None,
        "feeder_size": len(feeder) if feeder else 0,
        "notes_count": len(load_notes()),
        "data_dir_files": os.listdir(DATA_DIR),
        "last_error": err,
    })

if __name__ == "__main__":
    app.run(debug=True)
